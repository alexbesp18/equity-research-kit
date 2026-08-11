import openpyxl
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font

from er_kit.coveragebook import CoverageBook
from er_kit.visual_lint import SheetVisualContract, lint_workbook
from er_kit.visual_style import FONT_FAMILY


def _build_visual_book(tmp_path):
    out = tmp_path / "visual.xlsx"
    cb = CoverageBook(str(out))
    ws = cb.add_sheet("Decision", "0B3D2E")
    columns = [
        {"key": "ticker", "header": "Ticker", "width": 10, "type": "boldc"},
        {"key": "verdict", "header": "Verdict", "width": 14, "type": "verdict"},
        {"key": "price", "header": "Price", "width": 12, "type": "money"},
        {"key": "note", "header": "Note", "width": 42, "type": "text"},
    ]
    rows = [
        {
            "_currency": "USD",
            "ticker": "ALFA",
            "verdict": "buy_laggard",
            "price": 178.77,
            "note": "Clean wrapped operator note.",
        }
    ]
    start = cb.header_block(
        ws,
        len(columns),
        "Decision",
        "Visual QA surface",
        banner="Shared style tokens, locked panes, clean text wrapping.",
    )
    cb.table(ws, columns, rows, start_row=start, freeze=(start + 1, 2))
    cb.close()
    return out


def _issue_codes(report):
    return {issue.code for issue in report.issues}


def test_coveragebook_visual_contract_passes(tmp_path):
    out = _build_visual_book(tmp_path)
    report = lint_workbook(
        out,
        contracts=[SheetVisualContract("Decision", expected_freeze="C5")],
    )
    assert report.ok, [issue.message for issue in report.issues]

    ws = openpyxl.load_workbook(out)["Decision"]
    assert ws["A1"].font.name == FONT_FAMILY
    assert ws["A4"].font.name == FONT_FAMILY
    assert ws["A5"].font.name == FONT_FAMILY
    assert ws.sheet_view.showGridLines is False


def test_visual_lint_flags_random_font(tmp_path):
    out = _build_visual_book(tmp_path)
    wb = openpyxl.load_workbook(out)
    ws = wb["Decision"]
    ws["A1"].font = Font(name="Comic Sans MS", bold=True)
    wb.save(out)

    report = lint_workbook(out, contracts=[SheetVisualContract("Decision", expected_freeze="C5")])
    assert "random_font" in _issue_codes(report)


def test_visual_lint_flags_wrong_freeze(tmp_path):
    out = _build_visual_book(tmp_path)
    report = lint_workbook(out, contracts=[SheetVisualContract("Decision", expected_freeze="C6")])
    assert "freeze_mismatch" in _issue_codes(report)


def test_visual_lint_flags_unwrapped_clipping_risk(tmp_path):
    out = tmp_path / "clipping.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Bad"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 8
    ws["A1"] = "This operator note is too long for a narrow no-wrap cell."
    ws["A1"].font = Font(name=FONT_FAMILY)
    ws["A1"].alignment = Alignment(wrap_text=False)
    wb.save(out)

    report = lint_workbook(out, contracts=[SheetVisualContract("Bad")])
    assert "clipping_risk" in _issue_codes(report)


def test_visual_lint_flags_broad_conditional_format_range(tmp_path):
    out = tmp_path / "conditional.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Bad"
    ws.sheet_view.showGridLines = False
    ws["A1"] = 1
    ws["A1"].font = Font(name=FONT_FAMILY)
    ws.conditional_formatting.add(
        "A1:A1048576",
        CellIsRule(operator="greaterThan", formula=["0"]),
    )
    wb.save(out)

    report = lint_workbook(out, contracts=[SheetVisualContract("Bad")])
    assert "broad_conditional_format" in _issue_codes(report)


def test_visual_lint_flags_formula_error_literals(tmp_path):
    out = tmp_path / "formula-error.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Bad"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "#DIV/0!"
    ws["A1"].font = Font(name=FONT_FAMILY)
    wb.save(out)

    report = lint_workbook(out, contracts=[SheetVisualContract("Bad")])
    assert "formula_error" in _issue_codes(report)
