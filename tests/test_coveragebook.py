import openpyxl

from er_kit import universe as U
from er_kit.coveragebook import CoverageBook, compute_upside


def test_compute_upside_currency_safe():
    # merged record: USD ADR price with same-feed USD PT
    rec = {"fund": {"price": 30.55, "pt_mean": 34.04}}
    assert compute_upside(rec) == 11.4
    assert compute_upside({"fund": {}}) is None


def test_sheet_name_sanitized(tmp_path):
    out = tmp_path / "book.xlsx"
    cb = CoverageBook(str(out))
    ws = cb.add_sheet("Topic / Detail")  # contains illegal '/'
    cb.header_block(ws, 3, "T")
    cb.close()
    wb = openpyxl.load_workbook(str(out))
    assert "Topic - Detail" in wb.sheetnames


def test_build_decision_tab_opens_clean(tmp_path):
    out = tmp_path / "book.xlsx"
    cb = CoverageBook(str(out))
    ws = cb.add_sheet("Decision", "0B3D2E")
    cols = [
        {"key": "ticker", "header": "Ticker", "width": 9, "type": "boldc"},
        {"key": "verdict", "header": "Verdict", "width": 12, "type": "verdict"},
        {"key": "laggard", "header": "Laggard?", "width": 11, "type": "laggard"},
        {"key": "price", "header": "Price", "width": 10, "type": "money"},
        {"key": "upside", "header": "Upside", "width": 9, "type": "upside"},
        {"key": "note", "header": "Note", "width": 40, "type": "text", "max": 30},
    ]
    recs = [
        {
            "_currency": "USD",
            "ticker": "ALFA",
            "verdict": "buy_laggard",
            "laggard": "fair",
            "price": 45.97,
            "upside": 2.9,
            "note": "x" * 50,
        },
        {
            "_currency": "USD",
            "ticker": "BETA",
            "verdict": "pass",
            "laggard": "value_trap",
            "price": 33.63,
            "upside": 27.9,
            "note": "cheap but blew up",
        },
        # prose verdict should fall back to 'watch' without crashing
        {
            "_currency": "USD",
            "ticker": "WX",
            "verdict": "buy_laggard because ...",
            "laggard": None,
            "price": 10.0,
            "upside": None,
            "note": None,
        },
    ]
    start = cb.header_block(ws, len(cols), "Decision", "sub", banner="top-3 banner")
    last = cb.table(ws, cols, recs, start_row=start, freeze=(start + 1, 2))
    cb.close()
    assert last == start + 1 + len(recs)
    wb = openpyxl.load_workbook(str(out))
    d = wb["Decision"]
    # header is written at 0-indexed `start` => openpyxl row start+1; first data => start+2
    header_vals = [c.value for c in d[start + 1]]
    assert "Ticker" in header_vals
    first_data = [c.value for c in d[start + 2]]
    assert "ALFA" in first_data
    assert "buy_laggard" in first_data


def test_header_block_tracks_actual_table_column_count(tmp_path):
    # regression: stale header_block ncols must not bake a narrow banner before table() knows
    # the actual column count.
    out = tmp_path / "header-width.xlsx"
    cb = CoverageBook(str(out))
    ws = cb.add_sheet("D")
    start = cb.header_block(ws, 1, "Decision", banner="wide banner")
    cb.table(
        ws,
        [
            {"key": "a", "header": "A"},
            {"key": "b", "header": "B"},
            {"key": "c", "header": "C"},
        ],
        [{"a": 1, "b": 2, "c": 3}],
        start_row=start,
    )
    cb.close()
    ranges = {str(rng) for rng in openpyxl.load_workbook(str(out))["D"].merged_cells.ranges}
    assert "A1:C1" in ranges
    assert "A2:C2" in ranges


def test_merge_record_renders_non_usd_currency(tmp_path):
    # regression: table() must honor fund_currency (set by universe.merge), not only _currency,
    # else a JPY/EUR name renders a wrong '$' on the decision surface.
    out = tmp_path / "fx.xlsx"
    cb = CoverageBook(str(out))
    ws = cb.add_sheet("D")
    cols = [
        {"key": "ticker", "header": "Ticker", "width": 9, "type": "boldc"},
        {"key": "mc", "header": "Mkt cap", "width": 9, "type": "mcap"},
        {"key": "price", "header": "Price", "width": 9, "type": "money"},
    ]
    recs = [
        {"fund_currency": "JPY", "ticker": "BETA", "mc": 1.5e12, "price": 4320},  # no _currency
        {"fund_currency": "USD", "ticker": "ALFA", "mc": 2.7e10, "price": 178.77},
    ]
    cb.table(ws, cols, recs, start_row=0)
    cb.close()
    wb = openpyxl.load_workbook(str(out))
    d = wb["D"]
    # row 2 = fictional JPY listing -> mcap/price formats must NOT carry '$'
    assert "$" not in d.cell(2, 2).number_format
    assert "$" not in d.cell(2, 3).number_format
    # row 3 = fictional USD listing -> '$' present
    assert "$" in d.cell(3, 2).number_format
    # mcap magnitude: THREE trailing commas (/1e9 -> "B"), not the 2-comma /1e6 bug
    assert ",,," in d.cell(2, 2).number_format
    assert ",,," in d.cell(3, 2).number_format


def test_unknown_currency_uses_neutral_money_formats(tmp_path):
    # If no currency field exists at all, do not invent a false '$' format.
    out = tmp_path / "neutral.xlsx"
    cb = CoverageBook(str(out))
    ws = cb.add_sheet("D")
    cb.table(
        ws,
        [
            {"key": "price", "header": "Price", "type": "money"},
            {"key": "mc", "header": "Mkt cap", "type": "mcap"},
        ],
        [{"price": 12.34, "mc": 27.2e9}],
        start_row=0,
    )
    cb.close()
    d = openpyxl.load_workbook(str(out))["D"]
    assert "$" not in d.cell(2, 1).number_format
    assert "$" not in d.cell(2, 2).number_format
    assert ",,," in d.cell(2, 2).number_format


def test_non_dict_fund_field_does_not_crash(tmp_path):
    # a leaked non-dict 'fund' field must not crash the currency lookup mid-render
    out = tmp_path / "nf2.xlsx"
    cb = CoverageBook(str(out))
    ws = cb.add_sheet("D")
    cols = [
        {"key": "t", "header": "T", "type": "text"},
        {"key": "p", "header": "P", "type": "money"},
    ]
    cb.table(
        ws,
        cols,
        [{"t": "ALFA", "fund": "leaked string", "p": 10.0}, {"t": "BETA", "fund": ["x"], "p": 5.0}],
        start_row=0,
    )
    cb.close()  # must not raise
    assert openpyxl.load_workbook(str(out))["D"].max_row == 3
    assert compute_upside({"fund": "USD"}) is None  # helper guarded too


def test_laggard_strips_whitespace(tmp_path):
    out = tmp_path / "lg.xlsx"
    cb = CoverageBook(str(out))
    ws = cb.add_sheet("D")
    cb.table(
        ws,
        [{"key": "l", "header": "L", "type": "laggard"}],
        [{"l": " laggard "}, {"l": "junk"}],
        start_row=0,
    )
    cb.close()
    d = openpyxl.load_workbook(str(out))["D"]
    assert d.cell(2, 1).value == "laggard"  # padded valid bucket kept, not blanked
    assert d.cell(3, 1).value in (None, "")  # junk -> blank


def test_npint_upside_kept_and_colored(tmp_path):
    # numpy.int64 upside must be kept + colored (numbers.Real), like num/money columns
    import numpy as np

    out = tmp_path / "ui.xlsx"
    cb = CoverageBook(str(out))
    ws = cb.add_sheet("D")
    cb.table(
        ws, [{"key": "u", "header": "U", "type": "upside"}], [{"u": np.int64(15)}], start_row=0
    )
    cb.close()
    assert openpyxl.load_workbook(str(out))["D"].cell(2, 1).value == 15
    assert cb.upside_fmt(np.int64(20)) is not cb.f["num1"]  # colored band, not the fallback


def test_currency_none_falls_through(tmp_path):
    # an explicit _currency=None must fall through to USD ($), not be read as "not USD"
    out = tmp_path / "cn.xlsx"
    cb = CoverageBook(str(out))
    ws = cb.add_sheet("D")
    cb.table(
        ws,
        [{"key": "p", "header": "P", "type": "money"}],
        [{"p": 10.0, "_currency": None}],
        start_row=0,
    )
    cb.close()
    assert "$" in openpyxl.load_workbook(str(out))["D"].cell(2, 1).number_format


def test_exhaustive_coercion_no_crash(tmp_path):
    # numpy.datetime64 / ndarray / complex / range must NOT crash the build (whitelist->fallback)
    import numpy as np

    out = tmp_path / "ex.xlsx"
    cb = CoverageBook(str(out))
    ws = cb.add_sheet("D")
    cols = [
        {"key": "t", "header": "T", "type": "boldc"},
        {"key": "d", "header": "D", "type": "text"},
        {"key": "a", "header": "A", "type": "text"},
        {"key": "c", "header": "C", "type": "text"},
        {"key": "na", "header": "#An", "type": "center"},
    ]
    recs = [
        {
            "t": "ALFA",
            "d": np.datetime64("2026-02-15"),
            "a": np.array([1, 2, 3]),
            "c": complex(1, 2),
            "na": np.bool_(True),  # in a non-numeric cell -> coerced to text, not a crash
        }
    ]
    cb.table(ws, cols, recs, start_row=0)
    cb.close()  # must not raise
    d = openpyxl.load_workbook(str(out))["D"]
    assert "2026-02-15" in str(d.cell(2, 2).value)  # datetime64 rendered readable


def test_numpy_datetime64_time_renders_short_date(tmp_path):
    import numpy as np

    out = tmp_path / "ndt.xlsx"
    cb = CoverageBook(str(out))
    ws = cb.add_sheet("D")
    cb.table(
        ws,
        [{"key": "asof", "header": "As-of", "type": "text"}],
        [{"asof": np.datetime64("2026-02-15T16:30:45.123456789")}],
        start_row=0,
    )
    cb.close()
    assert openpyxl.load_workbook(str(out))["D"].cell(2, 1).value == "2026-02-15"


def test_npbool_and_string_blanked_in_numeric(tmp_path):
    import numpy as np

    out = tmp_path / "nb.xlsx"
    cb = CoverageBook(str(out))
    ws = cb.add_sheet("D")
    cols = [
        {"key": "u", "header": "U", "type": "upside"},
        {"key": "m", "header": "M", "type": "money"},
        {"key": "n", "header": "N", "type": "num"},
    ]
    cb.table(ws, cols, [{"u": np.bool_(True), "m": "NOTANUM", "n": np.int64(42)}], start_row=0)
    cb.close()
    d = openpyxl.load_workbook(str(out))["D"]
    assert d.cell(2, 1).value in (None, "")  # np.bool_ blanked
    assert d.cell(2, 2).value in (None, "")  # non-numeric string blanked
    assert d.cell(2, 3).value == 42  # legit numpy int kept


def test_nonfinite_blanked_in_text_cell_too(tmp_path):
    # NaN must blank in ANY cell type (n_analysts is often NaN and renders as center/text)
    out = tmp_path / "nt.xlsx"
    cb = CoverageBook(str(out))
    ws = cb.add_sheet("D")
    cols = [
        {"key": "na", "header": "#An", "type": "center"},
        {"key": "x", "header": "X", "type": "text"},
        {"key": "b", "header": "B", "type": "boldc"},
    ]
    cb.table(ws, cols, [{"na": float("nan"), "x": float("inf"), "b": float("-inf")}], start_row=0)
    cb.close()
    d = openpyxl.load_workbook(str(out))["D"]
    for col in (1, 2, 3):
        assert d.cell(2, col).value in (None, ""), f"col {col} not blanked"


def test_laggard_upside_bool_clamps(tmp_path):
    out = tmp_path / "cl.xlsx"
    cb = CoverageBook(str(out))
    ws = cb.add_sheet("D")
    cols = [
        {"key": "lag", "header": "L", "type": "laggard"},
        {"key": "u", "header": "U", "type": "upside"},
        {"key": "m", "header": "M", "type": "money"},
    ]
    # junk laggard -> '', non-numeric/bool in numeric cols -> blank
    cb.table(ws, cols, [{"lag": "not_an_enum", "u": "12%", "m": True}], start_row=0)
    cb.close()
    d = openpyxl.load_workbook(str(out))["D"]
    assert d.cell(2, 1).value in (None, "")  # junk laggard clamped
    assert d.cell(2, 2).value in (None, "")  # string upside coerced out
    assert d.cell(2, 3).value in (None, "")  # bool in money blanked


def test_nested_fund_currency_fallback(tmp_path):
    out = tmp_path / "nfc.xlsx"
    cb = CoverageBook(str(out))
    ws = cb.add_sheet("D")
    cols = [{"key": "m", "header": "M", "type": "mcap"}]
    # currency only nested under 'fund' -> must still pick the non-$ (JPY) format
    cb.table(ws, cols, [{"m": 1.5e12, "fund": {"currency": "JPY"}}], start_row=0)
    cb.close()
    assert "$" not in openpyxl.load_workbook(str(out))["D"].cell(2, 1).number_format


def test_datetime_cell_renders_as_date_string(tmp_path):
    import datetime

    out = tmp_path / "dt.xlsx"
    cb = CoverageBook(str(out))
    ws = cb.add_sheet("D")
    cols = [
        {"key": "t", "header": "T", "type": "boldc"},
        {"key": "asof", "header": "As-of", "type": "text"},
    ]
    cb.table(ws, cols, [{"t": "ALFA", "asof": datetime.datetime(2026, 6, 17, 16, 0)}], start_row=0)
    cb.close()
    c = openpyxl.load_workbook(str(out))["D"].cell(2, 2)
    assert c.value == "2026-06-17"  # ISO date string, not a 46190.66 serial


def test_nonfinite_numeric_cell_blanked(tmp_path):
    out = tmp_path / "nf.xlsx"
    cb = CoverageBook(str(out))
    ws = cb.add_sheet("D")
    cols = [
        {"key": "mc", "header": "MC", "type": "mcap"},
        {"key": "u", "header": "U", "type": "upside"},
        {"key": "p", "header": "P", "type": "money"},
    ]
    cb.table(ws, cols, [{"mc": float("nan"), "u": float("inf"), "p": float("nan")}], start_row=0)
    cb.close()
    d = openpyxl.load_workbook(str(out))["D"]
    for col in (1, 2, 3):
        v = d.cell(2, col).value
        assert v in (None, ""), f"col {col} rendered {v!r} (should be blank, not #NUM!)"


def test_bytes_cell_does_not_detonate_at_close(tmp_path):
    out = tmp_path / "by.xlsx"
    cb = CoverageBook(str(out))
    ws = cb.add_sheet("D")
    cb.table(ws, [{"key": "t", "header": "T", "type": "text"}], [{"t": b"binary"}], start_row=0)
    cb.close()  # must not raise at finalization
    assert openpyxl.load_workbook(str(out))["D"].cell(2, 1).value is not None


def test_add_sheet_handles_apostrophe_and_empty(tmp_path):
    # LLM-generated tab names with apostrophes / empties must not crash the build
    out = tmp_path / "ap.xlsx"
    cb = CoverageBook(str(out))
    cb.add_sheet("'AI' Accelerators")
    cb.add_sheet("Vera-Rubin 'rising intensity'")
    cb.add_sheet("")  # empty -> 'Sheet'
    cb.close()
    names = openpyxl.load_workbook(str(out)).sheetnames
    assert len(names) == 3 and len(set(names)) == 3
    assert not any(n.startswith("'") or n.endswith("'") for n in names)


def test_table_tolerates_non_dict_and_nonscalar(tmp_path):
    out = tmp_path / "rob.xlsx"
    cb = CoverageBook(str(out))
    ws = cb.add_sheet("D")
    cols = [
        {"key": "t", "header": "T", "type": "boldc"},
        {"key": "x", "header": "X", "type": "text"},
    ]
    # a None record (blank row) and a dict-valued cell (LLM leak) must not crash the render
    recs = [{"t": "A", "x": {"nested": 1}}, None, {"t": "C", "x": "ok"}]
    cb.table(ws, cols, recs, start_row=0)
    cb.close()
    d = openpyxl.load_workbook(str(out))["D"]
    assert d.max_row == 4  # header + 3 rows (None rendered as a blank row)
    assert "nested" in str(d.cell(2, 2).value)  # dict coerced to text


def test_verdict_cell_strips_whitespace(tmp_path):
    # renderer must reuse sanitize_verdict: a stray-space verdict should NOT fall back to 'watch'
    out = tmp_path / "v.xlsx"
    cb = CoverageBook(str(out))
    ws = cb.add_sheet("D")
    cols = [
        {"key": "t", "header": "T", "type": "boldc"},
        {"key": "v", "header": "V", "type": "verdict"},
    ]
    cb.table(ws, cols, [{"t": "A", "v": " buy_laggard"}, {"t": "B", "v": "junk"}], start_row=0)
    cb.close()
    d = openpyxl.load_workbook(str(out))["D"]
    assert d.cell(2, 2).value == "buy_laggard"  # stripped, not 'watch'
    assert d.cell(3, 2).value == "watch"  # genuine junk -> fallback


def test_sheet_name_collision_dedup(tmp_path):
    # long names sharing the first 31 chars must not crash the build (DuplicateWorksheetName)
    out = tmp_path / "dup.xlsx"
    cb = CoverageBook(str(out))
    n1 = "Datacenter Liquid Cooling - Cold Plate Single-Phase"
    n2 = "Datacenter Liquid Cooling - Cold Plate Two-Phase"
    cb.add_sheet(n1)
    cb.add_sheet(n2)
    cb.close()
    wb = openpyxl.load_workbook(str(out))
    assert len(wb.sheetnames) == 2
    assert len(set(wb.sheetnames)) == 2  # distinct


def test_upside_fmt_tolerates_non_numeric(tmp_path):
    out = tmp_path / "u.xlsx"
    cb = CoverageBook(str(out))
    ws = cb.add_sheet("D")
    cols = [
        {"key": "t", "header": "T", "type": "boldc"},
        {"key": "u", "header": "U", "type": "upside"},
    ]
    # None, NaN, and a stray string must not abort the render
    recs = [
        {"t": "A", "u": None},
        {"t": "B", "u": float("nan")},
        {"t": "C", "u": "N/A"},
        {"t": "D", "u": 14.0},
    ]
    cb.table(ws, cols, recs, start_row=0)
    cb.close()
    assert openpyxl.load_workbook(str(out))["D"].max_row == 5


def test_merge_duplicate_ticker_fails_loud():
    import pytest

    uni = [
        U.make_record("ALFA", "Alpha Example Corp.", "compute", tab="pure"),
        U.make_record(
            "ALFA", "Alpha Example Corp.", "networking", tab="pure"
        ),  # same ticker, 2 sub-verticals
    ]
    with pytest.raises(ValueError, match="duplicate ticker"):
        U.merge(uni)


def test_universe_merge():
    uni = [
        U.make_record("ALFA", "Alpha Example Corp.", "components", tab="pure"),
        U.make_record(
            "BETA", "Beta Example Ltd.", "infrastructure", tab="tangential", fund_currency="JPY"
        ),
    ]
    merged = U.merge(
        uni,
        fundamentals={"ALFA": {"price": 45.97, "currency": "USD", "pt_mean": 47.3}},
        dossiers={"ALFA": {"refreshed_verdict": "watch"}},
    )
    assert merged["ALFA"]["fund"]["price"] == 45.97
    assert merged["ALFA"]["fund_currency"] == "USD"
    assert merged["ALFA"]["dossier"]["refreshed_verdict"] == "watch"
    assert merged["BETA"]["fund"] == {}
    groups = U.by_subvertical(merged)
    assert groups["components"] == ["ALFA"]


def test_universe_merge_deep_copies_attached_layers():
    fund = {"price": 45.97, "currency": "USD", "nested": {"pt": 47.3}}
    tech = {"levels": ["support"]}
    dossier = {"analyst": {"recent_actions": [{"firm": "Example Research"}]}}
    foreign = {"line": {"ticker": "ALFA.F"}}
    uni = [U.make_record("ALFA", "Alpha Example Corp.", "components", prior={"notes": ["seed"]})]

    merged = U.merge(
        uni,
        fundamentals={"ALFA": fund},
        technicals={"ALFA": tech},
        foreign={"ALFA": foreign},
        dossiers={"ALFA": dossier},
    )
    rec = merged["ALFA"]
    rec["fund"]["nested"]["pt"] = 99
    rec["tech"]["levels"].append("resistance")
    rec["dossier"]["analyst"]["recent_actions"][0]["firm"] = "Changed"
    rec["foreign"]["line"]["ticker"] = "CHANGED"
    rec["prior"]["notes"].append("mutated")

    assert fund["nested"]["pt"] == 47.3
    assert tech["levels"] == ["support"]
    assert dossier["analyst"]["recent_actions"][0]["firm"] == "Example Research"
    assert foreign["line"]["ticker"] == "ALFA.F"
    assert uni[0]["prior"]["notes"] == ["seed"]
