import importlib.util
from pathlib import Path

import openpyxl


def _load_example_module():
    path = Path(__file__).resolve().parents[1] / "examples" / "build_coverage_example.py"
    spec = importlib.util.spec_from_file_location("build_coverage_example", path)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


def test_build_coverage_example_builds_workbook_and_calls_cleanopen(tmp_path, monkeypatch):
    example = _load_example_module()
    called = {}

    def fake_clean_open(path):
        called["path"] = path
        return True, "CLEAN OPEN OK | workbook=coverage_example.xlsx | sheets=2"

    monkeypatch.setattr(example, "clean_open", fake_clean_open)
    out, ok, msg = example.build_example(tmp_path / "coverage_example.xlsx")

    assert ok
    assert "CLEAN OPEN OK" in msg
    assert called["path"] == str(out)

    wb = openpyxl.load_workbook(str(out))
    assert wb.sheetnames == ["Decision", "Legend"]
    decision = wb["Decision"]
    assert decision.cell(5, 1).value == "ALFA"
    assert decision.cell(5, 3).value == "buy_laggard"
    assert decision.cell(6, 1).value == "BETA"
    assert decision.cell(6, 3).value == "watch"
    assert "$" in decision.cell(5, 5).number_format
    assert "$" not in decision.cell(6, 5).number_format
    assert "$" not in decision.cell(6, 6).number_format
    assert wb["Legend"].cell(2, 1).value == "verdict"
