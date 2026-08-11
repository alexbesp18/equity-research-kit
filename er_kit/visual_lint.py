"""Post-render visual contract checks for generated .xlsx coverage books."""

from __future__ import annotations

from collections.abc import Iterable
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from .visual_style import FONT_FAMILY

EXCEL_ERROR_VALUES = {
    "#DIV/0!",
    "#N/A",
    "#NAME?",
    "#NULL!",
    "#NUM!",
    "#REF!",
    "#VALUE!",
}


@dataclass(frozen=True)
class SheetVisualContract:
    """Optional expectations for a sheet after the workbook is written."""

    sheet_name: str
    expected_freeze: str | None = None
    require_gridlines_hidden: bool = True


@dataclass(frozen=True)
class VisualLintIssue:
    code: str
    sheet: str
    cell: str
    message: str


@dataclass
class VisualLintReport:
    path: Path
    issues: list[VisualLintIssue] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return not self.issues

    def assert_ok(self) -> None:
        if self.issues:
            details = "\n".join(f"{i.code} {i.sheet}!{i.cell}: {i.message}" for i in self.issues)
            raise AssertionError(f"visual lint failed for {self.path}:\n{details}")


def lint_workbook(
    path: str | Path,
    contracts: Iterable[SheetVisualContract] | None = None,
    allowed_fonts: set[str] | None = None,
    *,
    check_clipping: bool = True,
    check_conditional_formatting: bool = True,
    check_formula_errors: bool = True,
) -> VisualLintReport:
    """Inspect workbook XML for visual issues that unit tests otherwise miss."""

    xlsx_path = Path(path)
    allowed = allowed_fonts or {FONT_FAMILY}
    report = VisualLintReport(path=xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    value_wb = openpyxl.load_workbook(xlsx_path, data_only=True) if check_formula_errors else None
    contracts_by_sheet = {contract.sheet_name: contract for contract in contracts or []}

    seen_sheets: set[str] = set()
    for ws in wb.worksheets:
        seen_sheets.add(ws.title)
        used = _used_bounds(ws)
        contract = contracts_by_sheet.get(ws.title)
        if (
            contract
            and contract.require_gridlines_hidden
            and ws.sheet_view.showGridLines is not False
        ):
            report.issues.append(
                VisualLintIssue(
                    "gridlines_visible",
                    ws.title,
                    "sheet",
                    "gridlines must be hidden on operator-facing workbook tabs",
                )
            )
        if contract and contract.expected_freeze is not None:
            actual = _freeze_text(ws.freeze_panes)
            if actual != contract.expected_freeze:
                report.issues.append(
                    VisualLintIssue(
                        "freeze_mismatch",
                        ws.title,
                        "sheet",
                        f"expected freeze {contract.expected_freeze}, got {actual or 'none'}",
                    )
                )
        if used is None:
            continue
        min_row, max_row, min_col, max_col = used
        _check_fonts(report, ws, min_row, max_row, min_col, max_col, allowed)
        if check_clipping:
            _check_clipping(report, ws, min_row, max_row, min_col, max_col)
        if check_conditional_formatting:
            _check_conditional_formatting(report, ws, max_row, max_col)
        if check_formula_errors and value_wb is not None:
            _check_formula_errors(
                report, ws, value_wb[ws.title], min_row, max_row, min_col, max_col
            )
    for sheet_name in contracts_by_sheet:
        if sheet_name not in seen_sheets:
            report.issues.append(
                VisualLintIssue(
                    "missing_sheet",
                    sheet_name,
                    "sheet",
                    "sheet named in visual contract was not emitted",
                )
            )
    return report


def assert_visual_contract(
    path: str | Path,
    contracts: Iterable[SheetVisualContract] | None = None,
    allowed_fonts: set[str] | None = None,
) -> None:
    """Raise if lint_workbook finds visual contract failures."""

    lint_workbook(path, contracts=contracts, allowed_fonts=allowed_fonts).assert_ok()


def _used_bounds(ws) -> tuple[int, int, int, int] | None:
    min_row = min_col = 10**9
    max_row = max_col = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value in (None, ""):
                continue
            min_row = min(min_row, cell.row)
            max_row = max(max_row, cell.row)
            min_col = min(min_col, cell.column)
            max_col = max(max_col, cell.column)
    if max_row == 0:
        return None
    return min_row, max_row, min_col, max_col


def _freeze_text(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        return value
    return getattr(value, "coordinate", str(value))


def _check_fonts(
    report: VisualLintReport,
    ws,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
    allowed_fonts: set[str],
) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if cell.value in (None, ""):
                continue
            font_name = cell.font.name
            if font_name and font_name not in allowed_fonts:
                report.issues.append(
                    VisualLintIssue(
                        "random_font",
                        ws.title,
                        cell.coordinate,
                        f"font {font_name!r} is not in allowed fonts {sorted(allowed_fonts)}",
                    )
                )


def _check_clipping(
    report: VisualLintReport,
    ws,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            value = cell.value
            if not isinstance(value, str) or not value.strip():
                continue
            if value.startswith("=") or cell.alignment.wrap_text:
                continue
            width = _cell_width(ws, cell)
            longest_line = max(len(line) for line in value.splitlines())
            budget = max(12, int(width * 1.25))
            if longest_line > budget:
                report.issues.append(
                    VisualLintIssue(
                        "clipping_risk",
                        ws.title,
                        cell.coordinate,
                        f"text length {longest_line} exceeds no-wrap column budget {budget}",
                    )
                )


def _cell_width(ws, cell) -> float:
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return sum(
                _column_width(ws, get_column_letter(col_idx))
                for col_idx in range(merged_range.min_col, merged_range.max_col + 1)
            )
    return _column_width(ws, cell.column_letter)


def _column_width(ws, column_letter: str) -> float:
    width = ws.column_dimensions[column_letter].width
    if width is not None:
        return float(width)
    if ws.sheet_format.defaultColWidth:
        return float(ws.sheet_format.defaultColWidth)
    return 8.43


def _check_conditional_formatting(
    report: VisualLintReport,
    ws,
    used_max_row: int,
    used_max_col: int,
) -> None:
    for cf in ws.conditional_formatting:
        for cell_range in cf.sqref.ranges:
            range_text = str(cell_range)
            if cell_range.max_row >= 1_048_576 or cell_range.max_col >= 16_384:
                report.issues.append(
                    VisualLintIssue(
                        "broad_conditional_format",
                        ws.title,
                        range_text,
                        "conditional formatting must not target whole rows, whole columns, or whole sheets",
                    )
                )
            elif cell_range.max_row > used_max_row + 2000 or cell_range.max_col > used_max_col + 50:
                report.issues.append(
                    VisualLintIssue(
                        "loose_conditional_format",
                        ws.title,
                        range_text,
                        "conditional formatting range extends far beyond the used workbook surface",
                    )
                )


def _check_formula_errors(
    report: VisualLintReport,
    formula_ws,
    value_ws,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
) -> None:
    for row in formula_ws.iter_rows(
        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
    ):
        for cell in row:
            formula_value = cell.value
            cached_value = value_ws[cell.coordinate].value
            if formula_value in EXCEL_ERROR_VALUES or cached_value in EXCEL_ERROR_VALUES:
                report.issues.append(
                    VisualLintIssue(
                        "formula_error",
                        formula_ws.title,
                        cell.coordinate,
                        f"cell evaluates or renders as {formula_value or cached_value}",
                    )
                )
            elif isinstance(formula_value, str) and formula_value.startswith("="):
                literals = [err for err in EXCEL_ERROR_VALUES if err in formula_value]
                if literals:
                    report.issues.append(
                        VisualLintIssue(
                            "formula_error_literal",
                            formula_ws.title,
                            cell.coordinate,
                            f"formula contains Excel error literal {literals[0]}",
                        )
                    )
