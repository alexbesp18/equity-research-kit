"""Load a wide ticker-keyed technical workbook into per-ticker dictionaries.

The source file's prices are as-of its own `Updated` timestamp (often a prior close).
Use it for structure (RSI, trend, 52-week position, relative strength, drawdown, and
crosses), not as the authoritative live price.
"""

from __future__ import annotations

import openpyxl

# Sensible default columns for entry/structure analysis; override as needed.
DEFAULT_KEEP = [
    "Ticker",
    "Price",
    "Change%",
    "Updated",
    "MT_RSI_14",
    "MT_Entry_Score",
    "MT_Conviction",
    "LT_RSI_21",
    "LT_Price_vs_SMA200",
    "LT_SMA50_vs_SMA200",
    "LT_52W_Position",
    "LT_Trend",
    "LT_Months_in_Trend",
    "LT_Score",
    "AI_Key_Levels",
    "Cross_50_200",
    "Cross_50_200_Age",
    "Cross_10_50",
    "Cross_10_50_Age",
    "ST_ATR14_Pct",
    "MT_SMA50_Slope_20d",
    "LT_SMA200_Slope_50d",
    "LT_RangePos_252d",
    "LT_DD_52W",
    "LT_MaxDD_252d",
    "MT_RS_SPY_21d",
    "LT_RS_SPY_126d",
    "LT_RS_SPY_252d",
    "AvgDollarVol_20d",
    "AvgDollarVol_60d",
    "Market_Regime",
    "Market_VIX_Band",
    "AI_MT_Outlook",
    "AI_LT_Outlook",
    "AI_Risk_Level",
]


def load(path: str, tickers=None, keep=None, sheet: str | None = None) -> dict:
    """Return {ticker: {col: value}} for the requested tickers (or all).

    `tickers`: iterable to filter to (case-sensitive on the Ticker column); None = all.
    `keep`: columns to retain; None = DEFAULT_KEEP intersected with the file's headers.
    `sheet`: worksheet name; None = first sheet.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    # strip whitespace so a stray 'Ticker ' header still resolves
    hdr = [h.strip() if isinstance(h, str) else h for h in rows[0]]
    # structural ingestion -> fail loud on duplicate headers (silent last-wins would
    # corrupt RSI/trend/price reads if the upstream wide xlsx schema drifts)
    names = [h for h in hdr if h is not None]
    dups = sorted({h for h in names if names.count(h) > 1})
    if dups:
        raise ValueError(f"duplicate headers in {path}: {dups}")
    idx = {h: i for i, h in enumerate(hdr) if h is not None}
    if "Ticker" not in idx:
        raise ValueError(f"no 'Ticker' column in {path}; headers={hdr[:8]}")
    cols = [c for c in (keep or DEFAULT_KEEP) if c in idx]
    want = set(tickers) if tickers is not None else None
    out: dict = {}
    for r in rows[1:]:
        tval = r[idx["Ticker"]]
        if tval is None:
            continue
        tk = str(tval).strip()
        if not tk:
            continue  # skip whitespace-only ticker (junk '' key)
        if want is not None and tk not in want:
            continue
        # fail loud on duplicate ticker rows (same silent-data-loss class as dup headers)
        if tk in out:
            raise ValueError(f"duplicate ticker row in {path}: {tk!r}")
        row = {c: r[idx[c]] for c in cols}
        if "Ticker" in row:
            row["Ticker"] = tk
        out[tk] = row
    return out


def coverage(path: str, tickers, sheet: str | None = None) -> dict:
    """Which of `tickers` are present in the file vs missing."""
    have = set(load(path, tickers=tickers, keep=["Ticker"], sheet=sheet).keys())
    return {
        "covered": [t for t in tickers if t in have],
        "missing": [t for t in tickers if t not in have],
    }
